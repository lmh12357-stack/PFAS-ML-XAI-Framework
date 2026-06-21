import pandas as pd
import numpy as np
import os
import glob
import warnings
import xarray as xr
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font

warnings.filterwarnings('ignore')

# ================= 1. Global Configuration =================
INPUT_FILE = 'DATA.xlsx'
OUTPUT_FILE = 'POC_Imputed_Final.xlsx'

# Data Directories
DIR_POC = r'C:\Users\Administrator\PycharmProjects\PythonProject3\NASA_POC_Monthly_Data'
DIR_RRS = {
    488: r'C:\Users\Administrator\PycharmProjects\PythonProject3\NASA_Rrs488_Monthly_Data',
    547: r'C:\Users\Administrator\PycharmProjects\PythonProject3\NASA_Rrs547_Monthly_Data',
    645: r'C:\Users\Administrator\PycharmProjects\PythonProject3\NASA_Rrs645_Monthly_Data',
    678: r'C:\Users\Administrator\PycharmProjects\PythonProject3\NASA_Rrs678_Monthly_Data'
}

COASTAL_PADDING_LIMIT = 1  # 1 pixel padding (~55km) for coastal extrapolation

# Empirical Constants (Verbatim Restoration)
CI_SLOPE, CI_INTERCEPT = 171.30, 1.93
RATIO_SLOPE, RATIO_INTERCEPT = 1.78, 1.89

# Seamless Regional Geo-provinces
SEA_BOUNDARIES = {
    'Bohai': {'lat': (37.0, 42.0), 'lon': (117.0, 122.5)},
    'Yellow': {'lat': (31.0, 37.0), 'lon': (118.0, 127.0)},
    'East': {'lat': (23.0, 31.0), 'lon': (116.0, 131.0)},
    'South': {'lat': (3.0, 23.0), 'lon': (105.0, 121.0)}
}


# ================= 2. Core Engines =================

def get_region(lon, lat):
    for sea, bounds in SEA_BOUNDARIES.items():
        if bounds['lat'][0] <= lat <= bounds['lat'][1] and \
                bounds['lon'][0] <= lon <= bounds['lon'][1]:
            return sea
    return 'OpenOcean'


def build_temporal_index(directory, label="Dataset"):
    t_map = {}
    files = glob.glob(os.path.join(directory, "*.nc"))
    for f in files:
        parsed = False
        try:
            with xr.open_dataset(f) as ds:
                if 'time_coverage_start' in ds.attrs:
                    dt = pd.to_datetime(ds.attrs['time_coverage_start'])
                    t_map[(dt.year, dt.month)] = f
                    parsed = True
        except:
            pass
        if not parsed:
            try:
                fname = os.path.basename(f)
                year = int(fname[1:5])
                month = int(fname[5:7]) if fname[5:7].isdigit() else 1
                t_map[(year, month)] = f
            except:
                pass
    print(f"[SYSTEM] {label} Engine Mapped: {len(t_map)} temporal epochs.")
    return t_map


def batch_extract_bilinear(nc_path, target_coords, var_keyword):
    """
    High-Performance Batch Extraction with Bilinear Downscaling.
    target_coords: list of tuples [(idx, lat, lon), ...]
    Returns dict: {idx: val}
    """
    results = {idx: np.nan for idx, lat, lon in target_coords}

    if not nc_path or not os.path.exists(nc_path):
        return results

    try:
        with xr.open_dataset(nc_path) as ds:
            lat_key = 'lat' if 'lat' in ds.coords else 'latitude'
            lon_key = 'lon' if 'lon' in ds.coords else 'longitude'

            # Flexible keyword matching for varying NASA naming conventions
            var_name = next((k for k in ds.data_vars if var_keyword in k.lower() or (
                        var_keyword == 'rrs_645' and 'rrs' in k.lower() and '645' in k.lower())), None)
            if not var_name:
                var_name = next((k for k in ds.data_vars if 'rrs' in k.lower()), None)

            if not var_name:
                return results

            da = ds[var_name]

            if 'time' in da.dims:
                da = da.isel(time=0)
            if 'depth' in da.dims:
                da = da.isel(depth=0)

            # Sub-window Slicing
            lats = [c[1] for c in target_coords if pd.notna(c[1])]
            lons = [c[2] for c in target_coords if pd.notna(c[2])]
            if not lats or not lons: return results

            lat_min, lat_max = min(lats) - 2.0, max(lats) + 2.0
            lon_min, lon_max = min(lons) - 2.0, max(lons) + 2.0

            da_sub = da.sel({lat_key: slice(min(lat_min, lat_max), max(lat_min, lat_max)),
                             lon_key: slice(min(lon_min, lon_max), max(lon_min, lon_max))})

            if da_sub[lat_key].size == 0:
                da_sub = da.sel({lat_key: slice(max(lat_min, lat_max), min(lat_min, lat_max)),
                                 lon_key: slice(min(lon_min, lon_max), max(lon_min, lon_max))})

            # Extrapolate marine data to coastal margins
            da_padded = da_sub.ffill(dim=lon_key, limit=COASTAL_PADDING_LIMIT) \
                .bfill(dim=lon_key, limit=COASTAL_PADDING_LIMIT) \
                .ffill(dim=lat_key, limit=COASTAL_PADDING_LIMIT) \
                .bfill(dim=lat_key, limit=COASTAL_PADDING_LIMIT)

            # Batch Bilinear interpolation
            for idx, lat, lon in target_coords:
                if pd.isna(lat) or pd.isna(lon): continue
                try:
                    kwargs = {lat_key: lat, lon_key: lon, 'method': 'linear'}
                    val = float(da_padded.interp(**kwargs).values)
                    if not np.isnan(val) and val >= 0:
                        results[idx] = val
                except:
                    pass

            return results
    except Exception:
        return results


# ================= 3. Two-Pass Imputation Pipeline =================

def run_poc_imputation():
    print("\n--- INITIALIZING TEMPORAL ROUTING ENGINES ---")
    idx_poc = build_temporal_index(DIR_POC, "NASA_POC")
    idx_rrs = {w: build_temporal_index(path, f"NASA_Rrs{w}") for w, path in DIR_RRS.items()}

    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(f"[CRITICAL ERROR] Missing source file: {INPUT_FILE}")

    df = pd.read_excel(INPUT_FILE)
    df.columns = df.columns.str.strip()

    col_target = 'POC (mg/m3)'
    col_source = 'POC_Source'

    # [CRITICAL] 100% Reset logic for confirmed missing data
    df[col_target] = np.nan
    df[col_source] = 'Unfilled'

    df['Temp_Region'] = df.apply(lambda r: get_region(r['Longitude'], r['Latitude']), axis=1)
    df['Temp_Time'] = pd.to_datetime(df['Samplingtime'], errors='coerce')
    df['Temp_Month'] = df['Temp_Time'].dt.month
    df['Temp_Year'] = df['Temp_Time'].dt.year

    audit = {'L1_POC_Bilinear': 0, 'L2_Rrs_Bilinear_Inverted': 0, 'L3_Median': 0}

    # --- PHASE 1: Physical Inversion (L1 & L2) via I/O Optimized Batching ---
    print("\n--- PHASE 1: Physical Extraction & Multi-Band Optical Inversion ---")

    unfilled_df = df[df[col_source] == 'Unfilled']
    grouped = unfilled_df.groupby(['Temp_Year', 'Temp_Month'])

    for (year, month), group in grouped:
        if pd.isna(year) or pd.isna(month): continue

        target_coords = [(idx, row['Latitude'], row['Longitude']) for idx, row in group.iterrows()]

        # Level 1: NASA POC Direct Bilinear Downscaling
        f_l1 = idx_poc.get((int(year), int(month)))
        l1_results = batch_extract_bilinear(f_l1, target_coords, 'poc')

        remaining_coords_for_l2 = []
        for idx in group.index:
            val_l1 = l1_results.get(idx, np.nan)
            if pd.notna(val_l1):
                df.at[idx, col_target] = val_l1
                df.at[idx, col_source] = "NASA_Bilinear_Downscaled"
                audit['L1_POC_Bilinear'] += 1
            else:
                remaining_coords_for_l2.append((idx, df.at[idx, 'Latitude'], df.at[idx, 'Longitude']))

        # Level 2: Rrs Multi-Band Inversion
        if remaining_coords_for_l2:
            rrs_results = {w: {} for w in [488, 547, 645, 678]}
            for w in [488, 547, 645, 678]:
                f_rrs = idx_rrs[w].get((int(year), int(month)))
                if f_rrs:
                    rrs_results[w] = batch_extract_bilinear(f_rrs, remaining_coords_for_l2,
                                                            f'rrs_{w}' if w not in [645, 678] else 'rrs')

            for idx, lat, lon in remaining_coords_for_l2:
                rrs_vals = {w: rrs_results[w].get(idx, np.nan) for w in [488, 547, 645, 678]}

                # If all 4 bands were successfully extracted via bilinear interpolation
                if all(pd.notna(v) for v in rrs_vals.values()):
                    try:
                        if rrs_vals[488] >= rrs_vals[547]:
                            w_slope = (547 - 488) / (678 - 488)
                            baseline = rrs_vals[488] + w_slope * (rrs_vals[678] - rrs_vals[488])
                            CI = rrs_vals[547] - baseline
                            log_poc = CI_SLOPE * CI + CI_INTERCEPT
                        else:
                            X = rrs_vals[645] / rrs_vals[547]
                            log_poc = RATIO_SLOPE * X + RATIO_INTERCEPT

                        df.at[idx, col_target] = 10 ** log_poc  # Full precision
                        df.at[idx, col_source] = "Rrs_Bilinear_Inverted"
                        audit['L2_Rrs_Bilinear_Inverted'] += 1
                    except:
                        pass

    # --- PHASE 2: Dynamic Median Imputation (L3) ---
    print("\n--- PHASE 2: Dynamic Regional Median Imputation ---")
    unfilled = df[df[col_source] == 'Unfilled'].index
    for i in unfilled:
        reg, mo = df.at[i, 'Temp_Region'], df.at[i, 'Temp_Month']

        # Priority 1: Current Month + Current Region (Confirmed L1/L2 Pool)
        mask_p1 = (df['Temp_Region'] == reg) & (df['Temp_Month'] == mo) & (df[col_target].notna())
        val_l3 = df.loc[mask_p1, col_target].median()

        # Priority 2: All Months + Current Region
        if pd.isna(val_l3):
            mask_p2 = (df['Temp_Region'] == reg) & (df[col_target].notna())
            val_l3 = df.loc[mask_p2, col_target].median()

        # Priority 3: Global Base
        if pd.isna(val_l3):
            val_l3 = df[col_target].median()

        df.at[i, col_target] = max(0.0, float(val_l3)) if pd.notna(val_l3) else 0.0  # Full precision
        df.at[i, col_source] = "Regional_Median"
        audit['L3_Median'] += 1

    df.drop(columns=['Temp_Region', 'Temp_Time', 'Temp_Month', 'Temp_Year'], inplace=True, errors='ignore')

    # ================= 4. Data Export & Highlight =================
    df.to_excel(OUTPUT_FILE, index=False)

    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        r_font = Font(color="FF0000")
        h_map = {c.value: idx + 1 for idx, c in enumerate(ws[1])}
        if col_target in h_map and col_source in h_map:
            t_idx, s_idx = h_map[col_target], h_map[col_source]
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=s_idx).value != "Original":
                    ws.cell(row=r, column=t_idx).font = r_font
                    ws.cell(row=r, column=t_idx).number_format = '0.00'  # Presentation Layer Truncation ONLY
        wb.save(OUTPUT_FILE)
    except Exception:
        pass

    # ================= 5. Full-Spectrum Audit =================
    print("\n" + "=" * 65 + "\n      NASA-POC BILINEAR DOWNSCALING AUDIT\n" + "=" * 65)
    print(f"Level 1 - NASA_POC Direct Bilinear       : {audit['L1_POC_Bilinear']}")
    print(f"Level 2 - Rrs_Multi-Band Bilinear Invert : {audit['L2_Rrs_Bilinear_Inverted']}")
    print(f"Level 3 - Regional_Median (Dynamic Pool) : {audit['L3_Median']}")
    print("=" * 65)


if __name__ == "__main__":
    run_poc_imputation()