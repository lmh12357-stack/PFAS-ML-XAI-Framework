import pandas as pd
import numpy as np
import os
import glob
import warnings
import xarray as xr
from sklearn.neighbors import BallTree, KNeighborsRegressor
from openpyxl import load_workbook
from openpyxl.styles import Font

warnings.filterwarnings('ignore')

# ================= 1. Global Configuration =================
INPUT_EXCEL = "DATA.xlsx"
NC_DIR = "NASA_Chl-a_Monthly_Data"
OUTPUT_FILE = "Chla_Final_Cleaned.xlsx"

COASTAL_PADDING_LIMIT = 1  # 1 pixel padding (~55km for 0.5 deg grid) for coastal extrapolation

# Spatial Constraint Parameters for Downstream L2
SEARCH_RADIUS_L2_KNN = 0.5

# Haversine Conversion Constants
EARTH_RADIUS_KM = 6371.01
DEG_TO_KM = 111.32
SEARCH_RADIUS_RAD = (0.5 * DEG_TO_KM) / EARTH_RADIUS_KM

# Seamless Regional Geo-provinces
SEA_BOUNDARIES = {
    'Bohai': {'lat': (37.0, 42.0), 'lon': (117.0, 122.5)},
    'Yellow': {'lat': (31.0, 37.0), 'lon': (118.0, 127.0)},
    'East': {'lat': (23.0, 31.0), 'lon': (116.0, 131.0)},
    'South': {'lat': (3.0, 23.0), 'lon': (105.0, 121.0)}
}


# ================= 2. Core Methodologies =================

def get_region(lon, lat):
    for sea, bounds in SEA_BOUNDARIES.items():
        if bounds['lat'][0] <= lat <= bounds['lat'][1] and \
                bounds['lon'][0] <= lon <= bounds['lon'][1]:
            return sea
    return 'OpenOcean'


def build_temporal_index(nc_dir):
    temporal_map = {}
    nc_files = glob.glob(os.path.join(nc_dir, "*.nc"))
    for file_path in nc_files:
        parsed = False
        try:
            with xr.open_dataset(file_path) as dataset:
                if 'time_coverage_start' in dataset.attrs:
                    dt_obj = pd.to_datetime(dataset.attrs['time_coverage_start'])
                    temporal_map[(dt_obj.year, dt_obj.month)] = file_path
                    parsed = True
        except:
            pass
        if not parsed:
            try:
                fname = os.path.basename(file_path)
                year, month = int(fname[1:5]), int(fname[5:7]) if fname[5:7].isdigit() else 1
                temporal_map[(year, month)] = file_path
            except:
                pass
    print(f"[SYSTEM] Temporal Engine Mapped: {len(temporal_map)} epochs.")
    return temporal_map


def batch_extract_chla_bilinear(nc_path, target_coords):
    """
    Tier 1: High-Performance Batch Extraction.
    Uses sub-window slicing to eliminate global matrix overhead, followed by Bilinear Downscaling.
    target_coords: list of tuples [(idx, lat, lon), ...]
    Returns dict: {idx: val}
    """
    results = {idx: np.nan for idx, lat, lon in target_coords}

    if not nc_path or not os.path.exists(nc_path):
        return results

    try:
        with xr.open_dataset(nc_path) as ds:
            lat_key = 'lat' if 'lat' in ds.coords else 'latitude'
            lon_key = 'lon' if 'lon' in ds.coords else 'longitude'
            var_name = next((k for k in ds.data_vars if 'chl' in k.lower()), None)

            if not var_name:
                return results

            da = ds[var_name]

            if 'time' in da.dims:
                da = da.isel(time=0)
            if 'depth' in da.dims:
                da = da.isel(depth=0)

            # --- Sub-window Slicing Optimization ---
            lats = [c[1] for c in target_coords if pd.notna(c[1])]
            lons = [c[2] for c in target_coords if pd.notna(c[2])]

            if not lats or not lons:
                return results

            lat_min, lat_max = min(lats) - 2.0, max(lats) + 2.0
            lon_min, lon_max = min(lons) - 2.0, max(lons) + 2.0

            da_sub = da.sel({lat_key: slice(min(lat_min, lat_max), max(lat_min, lat_max)),
                             lon_key: slice(min(lon_min, lon_max), max(lon_min, lon_max))})

            # Re-check orientation to ensure correct slicing
            if da_sub[lat_key].size == 0:
                da_sub = da.sel({lat_key: slice(max(lat_min, lat_max), min(lat_min, lat_max)),
                                 lon_key: slice(min(lon_min, lon_max), max(lon_min, lon_max))})

            # Extrapolate marine data to coastal margins within the sub-window
            da_padded = da_sub.ffill(dim=lon_key, limit=COASTAL_PADDING_LIMIT) \
                .bfill(dim=lon_key, limit=COASTAL_PADDING_LIMIT) \
                .ffill(dim=lat_key, limit=COASTAL_PADDING_LIMIT) \
                .bfill(dim=lat_key, limit=COASTAL_PADDING_LIMIT)

            # Batch Bilinear interpolation
            for idx, lat, lon in target_coords:
                if pd.isna(lat) or pd.isna(lon): continue

                try:
                    kwargs = {lat_key: lat, lon_key: lon, 'method': 'linear'}
                    val = float(da_padded.interp(**kwargs).values)
                    if not np.isnan(val) and val >= 0:
                        results[idx] = val
                except:
                    pass

            return results
    except Exception:
        return results


# ================= 3. Two-Pass Sequential Assimilation Pipeline =================

def run_chla_imputation():
    if not os.path.exists(INPUT_EXCEL):
        raise FileNotFoundError(f"[CRITICAL ERROR] Missing source file: {INPUT_EXCEL}")

    temporal_index = build_temporal_index(NC_DIR)
    if not temporal_index:
        print(f"[WARNING] No valid .nc files found in {NC_DIR}. Proceeding with L2/L3 only.")

    df = pd.read_excel(INPUT_EXCEL)
    df.columns = df.columns.str.strip()

    col_target = 'Chla' if 'Chla' in df.columns else 'Chl-a'
    col_source = f'{col_target}_Source'

    # [CRITICAL] State Initialization: Strict alignment with Original Data
    if col_source not in df.columns:
        df.insert(df.columns.get_loc(col_target) + 1, col_source, np.nan)

    mask_original = df[col_target].notna() & (df.get(col_source, 'Unfilled') != 'NASA_Bilinear_Downscaled') & (
                df.get(col_source, 'Unfilled') != 'Climatological_Mean') & (
                                df.get(col_source, 'Unfilled') != 'Spatial_KNN') & (
                                df.get(col_source, 'Unfilled') != 'Regional_Median')

    df[col_source] = 'Unfilled'
    df.loc[mask_original, col_source] = 'Original Data'

    df['Temp_Region'] = df.apply(lambda row: get_region(row['Longitude'], row['Latitude']), axis=1)
    df['Temp_Time'] = pd.to_datetime(df['Samplingtime'], errors='coerce')
    df['Temp_Month'] = df['Temp_Time'].dt.month
    df['Temp_Year'] = df['Temp_Time'].dt.year

    audit = {'L1': 0, 'L1.5': 0, 'L2': 0, 'L3': 0}
    diag_l3 = []

    # --- PHASE 1: Physical Acquisition (L1 & L1.5) via I/O Optimized Batching ---
    print("\n--- PHASE 1: Executing NASA Satellite Bilinear Downscaling (Batch Mode) ---")

    unfilled_df = df[df[col_source] == 'Unfilled']
    grouped = unfilled_df.groupby(['Temp_Year', 'Temp_Month'])

    for (year, month), group in grouped:
        if pd.isna(year) or pd.isna(month): continue

        target_coords = [(idx, row['Latitude'], row['Longitude']) for idx, row in group.iterrows()]

        # Tier 1: Primary Satellite Extraction (Bilinear)
        nc_file = temporal_index.get((int(year), int(month)))
        l1_results = batch_extract_chla_bilinear(nc_file, target_coords)

        # Apply L1 results and identify remaining gaps for L1.5
        remaining_coords = []
        for idx in group.index:
            val_l1 = l1_results.get(idx, np.nan)
            if pd.notna(val_l1):
                df.at[idx, col_target] = val_l1
                df.at[idx, col_source] = "NASA_Bilinear_Downscaled"
                audit['L1'] += 1
            else:
                remaining_coords.append((idx, df.at[idx, 'Latitude'], df.at[idx, 'Longitude']))

        # Tier 1.5: Climatological Substitution (Bilinear)
        if remaining_coords:
            clim_files = [path for (y, m), path in temporal_index.items() if m == int(month)]

            # Aggregate climatological results across all available years for this month
            clim_aggregates = {idx: [] for idx, _, _ in remaining_coords}
            for cf in clim_files:
                cf_results = batch_extract_chla_bilinear(cf, remaining_coords)
                for idx, val in cf_results.items():
                    if pd.notna(val):
                        clim_aggregates[idx].append(val)

            for idx, vals in clim_aggregates.items():
                if vals:
                    df.at[idx, col_target] = np.mean(vals)
                    df.at[idx, col_source] = "Climatological_Mean"
                    audit['L1.5'] += 1

    # --- [RELOAD] Dynamic Background Field Update ---
    valid_pool = df[df[col_target].notna()]
    ref_coords_rad = np.radians(valid_pool[['Latitude', 'Longitude']].values)
    ref_vals = valid_pool[col_target].values

    # --- PHASE 2: Statistical Imputation (L2 & L3) ---
    print("\n--- PHASE 2: Executing Topologic KNN & Sequential Regional Median ---")
    unfilled_indices = df[df[col_source] == 'Unfilled'].index

    for idx in unfilled_indices:
        lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
        region = df.at[idx, 'Temp_Region']

        # Tier 2: Spatial KNN (Using Haversine Engine)
        l2_processed = False
        if len(ref_coords_rad) >= 3 and pd.notna(lat) and pd.notna(lon):
            tree = BallTree(ref_coords_rad, metric='haversine')
            query_rad = np.radians([[lat, lon]])
            nb_indices = tree.query_radius(query_rad, r=SEARCH_RADIUS_RAD)[0]

            if len(nb_indices) >= 3:
                knn = KNeighborsRegressor(
                    n_neighbors=min(5, len(nb_indices)),
                    weights='distance',
                    algorithm='ball_tree',
                    metric='haversine'
                )
                knn.fit(ref_coords_rad[nb_indices], ref_vals[nb_indices])
                val_l2 = float(knn.predict(query_rad)[0])

                df.at[idx, col_target] = max(0.0, val_l2)
                df.at[idx, col_source] = "Spatial_KNN"
                audit['L2'] += 1
                l2_processed = True

        if l2_processed: continue

        # Tier 3: Dynamic Regional Median (Priority: Region+Month -> Region -> Global)
        target_mo = df.at[idx, 'Temp_Month']
        mask_rm = (df['Temp_Region'] == region) & (df['Temp_Month'] == target_mo) & (df[col_target].notna())
        val_l3 = df.loc[mask_rm, col_target].median()

        if pd.isna(val_l3):
            mask_r = (df['Temp_Region'] == region) & (df[col_target].notna())
            val_l3 = df.loc[mask_r, col_target].median()

        if pd.isna(val_l3):
            val_l3 = np.median(ref_vals)

        df.at[idx, col_target] = max(0.0, float(val_l3)) if pd.notna(val_l3) else 0.0
        df.at[idx, col_source] = "Regional_Median"
        audit['L3'] += 1
        diag_l3.append((lon, lat, region))

    df.drop(columns=['Temp_Region', 'Temp_Time', 'Temp_Month', 'Temp_Year'], inplace=True, errors='ignore')

    # ================= 4. Data Export & Formatting =================
    df.to_excel(OUTPUT_FILE, index=False)

    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        r_font = Font(color="FF0000")
        h_map = {c.value: i + 1 for i, c in enumerate(ws[1])}

        if col_target in h_map and col_source in h_map:
            t_idx, s_idx = h_map[col_target], h_map[col_source]
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=s_idx).value != "Original Data":
                    ws.cell(row=r, column=t_idx).font = r_font
                    ws.cell(row=r, column=t_idx).number_format = '0.00'
        wb.save(OUTPUT_FILE)
    except Exception:
        pass

    # ================= 5. Audit Report =================
    print("\n" + "=" * 55 + "\n      Chl-a BILINEAR DOWNSCALING AUDIT\n" + "=" * 55)
    print(f"L1 - NASA Bilinear Downscaled     : {audit['L1']}")
    print(f"L1.5 - Climatological (Bilinear)  : {audit['L1.5']}")
    print(f"L2 - Spatial_KNN (Enriched Pool)  : {audit['L2']}")
    print(f"L3 - Regional_Median              : {audit['L3']}")
    print("=" * 55)


if __name__ == "__main__":
    run_chla_imputation()