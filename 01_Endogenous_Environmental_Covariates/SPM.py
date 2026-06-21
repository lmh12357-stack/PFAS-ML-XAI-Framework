import pandas as pd
import numpy as np
import os
import glob
import xarray as xr
import warnings
from sklearn.neighbors import BallTree
from pykrige.ok import OrdinaryKriging
from openpyxl import load_workbook
from openpyxl.styles import Font

warnings.filterwarnings('ignore')

# ================= 1. Global Academic Configuration =================
if not os.path.exists('DATA.xlsx'):
    raise FileNotFoundError("[CRITICAL ERROR] Source file DATA.xlsx not found.")

INPUT_FILE = 'DATA.xlsx'
OUTPUT_FILE = 'DATA_SPM_Imputed_Final.xlsx'
NC_DIR = r"C:\Users\Administrator\PycharmProjects\PythonProject3\CEMES_SPM"

SEARCH_RADIUS_DEG = 0.5
COASTAL_PADDING_LIMIT = 1  # 1 pixel padding (~55km for 0.5 deg grid) for coastal extrapolation

# Haversine Conversion Constants
EARTH_RADIUS_KM = 6371.01
DEG_TO_KM = 111.32
SEARCH_RADIUS_RAD = (SEARCH_RADIUS_DEG * DEG_TO_KM) / EARTH_RADIUS_KM


# ================= 2. Core Robust Extraction Modules =================

def extract_nc_spm_bilinear(ds, lat, lon, time):
    """
    Tier 1: SPM extraction featuring Bilinear Downscaling for sub-grid mapping.
    Includes coastal padding to prevent NaN mask blindness in nearshore zones.
    """
    try:
        ds_time = ds['SPM'].sel(time=time, method='nearest')

        # Extrapolate marine data to coastal margins to prevent interpolation failure
        da_padded = ds_time.ffill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .ffill(dim='latitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='latitude', limit=COASTAL_PADDING_LIMIT)

        # Bilinear interpolation for smooth sub-grid spatial gradient recovery
        val = float(da_padded.interp(latitude=lat, longitude=lon, method='linear').values)
        return val if not np.isnan(val) and val >= 0 else np.nan
    except Exception:
        return np.nan


def extract_clim_spm_bilinear(ds_clim, lat, lon, month):
    """
    Tier 3: Climatological extraction using Bilinear Downscaling.
    """
    try:
        ds_month = ds_clim['SPM'].sel(month=month)

        da_padded = ds_month.ffill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .ffill(dim='latitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='latitude', limit=COASTAL_PADDING_LIMIT)

        val = float(da_padded.interp(latitude=lat, longitude=lon, method='linear').values)
        return val if not np.isnan(val) and val >= 0 else np.nan
    except Exception:
        return np.nan


# ================= 3. Main Sequential Assimilation Pipeline =================

def run_spm_imputation():
    print("[SYSTEM] Initializing SPM Sequential Assimilation Pipeline...")

    nc_files = glob.glob(os.path.join(NC_DIR, "*.nc"))
    if nc_files:
        print(f"[SYSTEM] Loading {len(nc_files)} CMEMS netCDF files into memory...")
        ds_spm = xr.open_mfdataset(nc_files, combine='by_coords')

        if 'time' in ds_spm.indexes and not ds_spm.indexes['time'].is_unique:
            ds_spm = ds_spm.sel(time=~ds_spm.get_index('time').duplicated())

        ds_spm.load()

        ds_clim = ds_spm.groupby('time.month').mean('time')
        print("[SYSTEM] CMEMS Hypercube & Climatological Background Engine Active.")
    else:
        ds_spm = None
        ds_clim = None
        print("[WARNING] CMEMS Data not found. Pipeline will bypass satellite imputation.")

    all_sheets = pd.read_excel(INPUT_FILE, sheet_name=None)
    audit_stats = {'L1_Success': 0, 'L2_Local_Kriging': 0, 'L3_Climatology': 0}
    imputed_records = []

    for sheet_name, df in all_sheets.items():
        print(f"\n--- Processing Sheet: {sheet_name} ---")
        df.columns = [str(c).strip() for c in df.columns]

        if 'SPM (mg/L)' in df.columns:
            df.rename(columns={'SPM (mg/L)': 'SPM'}, inplace=True)
        if 'SPM' not in df.columns:
            df['SPM'] = np.nan

        if 'SPM_Source' not in df.columns:
            spm_idx = df.columns.get_loc('SPM')
            df.insert(spm_idx + 1, 'SPM_Source', np.nan)

        df['SPM'] = pd.to_numeric(df['SPM'], errors='coerce')

        if 'Samplingtime' in df.columns:
            df['Temp_Time'] = pd.to_datetime(df['Samplingtime'].astype(str).str.replace('.', '-'), errors='coerce')
            df['Temp_YM'] = df['Temp_Time'].dt.to_period('M')
            df['Temp_Month'] = df['Temp_Time'].dt.month
        else:
            df['Temp_Time'] = pd.NaT

        mask_original = df['SPM'].notna() & (df['SPM_Source'] != 'CMEMS_Bilinear_Downscaled') & (
                    df['SPM_Source'] != 'Local_Kriging') & (df['SPM_Source'] != 'Climatological_Bilinear_Fallback')

        df['SPM_Source'] = 'Unfilled'
        df.loc[mask_original, 'SPM_Source'] = 'Original Data'

        if ds_spm is not None:
            # --- STAGE 1: CMEMS Extraction (Bilinear) ---
            missing_idx = df[
                df['SPM'].isna() & df['Temp_Time'].notna() & df['Latitude'].notna() & df['Longitude'].notna()].index

            for idx in missing_idx:
                lat, lon, time = df.loc[idx, ['Latitude', 'Longitude', 'Temp_Time']]

                val_sat = extract_nc_spm_bilinear(ds_spm, lat, lon, time)
                if pd.notna(val_sat) and val_sat >= 0:
                    df.at[idx, 'SPM'] = val_sat
                    df.at[idx, 'SPM_Source'] = 'CMEMS_Bilinear_Downscaled'
                    audit_stats['L1_Success'] += 1
                    imputed_records.append((sheet_name, idx))

            # --- STAGE 2: Local Ordinary Kriging (Assimilated Pool 1) ---
            if 'Temp_YM' in df.columns:
                grouped_ym = df.dropna(subset=['Temp_YM']).groupby('Temp_YM')
                for ym, group in grouped_ym:
                    pool_1 = group[group['SPM'].notna()]
                    if len(pool_1) < 3: continue

                    # Haversine Enforcement: Extract as [Lat, Lon] and convert to radians
                    ref_coords_raw = pool_1[['Latitude', 'Longitude']].values
                    ref_coords_rad = np.radians(ref_coords_raw)
                    ref_vals = pool_1['SPM'].values

                    tree_pool = BallTree(ref_coords_rad, metric='haversine')

                    unfilled_idx = group[group['SPM'].isna()].index
                    for idx in unfilled_idx:
                        lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
                        if pd.isna(lat) or pd.isna(lon): continue

                        query_rad = np.radians([[lat, lon]])
                        nb_idx = tree_pool.query_radius(query_rad, r=SEARCH_RADIUS_RAD)[0]

                        if len(nb_idx) >= 3:
                            try:
                                # OrdinaryKriging requires [lon, lat] for (x, y).
                                ok = OrdinaryKriging(
                                    ref_coords_raw[nb_idx, 1], ref_coords_raw[nb_idx, 0], ref_vals[nb_idx],
                                    variogram_model='linear', verbose=False, enable_plotting=False,
                                    coordinates_type='geographic'
                                )
                                z, _ = ok.execute('grid', [lon], [lat])
                                val_krig = float(z.data[0][0])

                                if pd.notna(val_krig) and val_krig >= 0:
                                    df.at[idx, 'SPM'] = val_krig
                                    df.at[idx, 'SPM_Source'] = 'Local_Kriging'
                                    audit_stats['L2_Local_Kriging'] += 1
                                    imputed_records.append((sheet_name, idx))
                            except Exception:
                                pass

            # --- STAGE 3: Pixel-wise Climatology Fallback (Bilinear) ---
            unfilled_idx_p3 = df[
                df['SPM'].isna() & df['Latitude'].notna() & df['Longitude'].notna() & df['Temp_Month'].notna()].index
            for idx in unfilled_idx_p3:
                lat, lon, month = df.at[idx, 'Latitude'], df.at[idx, 'Longitude'], df.at[idx, 'Temp_Month']

                val_clim = extract_clim_spm_bilinear(ds_clim, lat, lon, month)
                if pd.notna(val_clim) and val_clim >= 0:
                    df.at[idx, 'SPM'] = val_clim
                    df.at[idx, 'SPM_Source'] = 'Climatological_Bilinear_Fallback'
                    audit_stats['L3_Climatology'] += 1
                    imputed_records.append((sheet_name, idx))

        df.drop(columns=['Temp_Time', 'Temp_YM', 'Temp_Month'], inplace=True, errors='ignore')
        all_sheets[sheet_name] = df

    # ================= 4. Data Export & Formatting =================
    print(f"\n[SYSTEM] Executing IO Writing to {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    try:
        wb = load_workbook(OUTPUT_FILE)
        red_font = Font(color="FF0000")

        for sheet_name in all_sheets.keys():
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]

            header_map = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
            spm_col_idx = header_map.get('SPM')

            if spm_col_idx:
                for sn, row_idx in imputed_records:
                    if sn == sheet_name:
                        ws.cell(row=row_idx + 2, column=spm_col_idx).font = red_font
                        ws.cell(row=row_idx + 2, column=spm_col_idx).number_format = '0.00'  # Formatting Layer Only
        wb.save(OUTPUT_FILE)
    except Exception as e:
        print(f"[WARNING] Style formatting exception: {e}")

    # ================= 5. Academic Audit Report =================
    print("\n" + "=" * 55)
    print("      SPM BILINEAR DOWNSCALING ASSIMILATION AUDIT")
    print("=" * 55)
    print(f"Level 1 - CMEMS Bilinear Downscaled   : {audit_stats['L1_Success']}")
    print(f"Level 2 - Local Kriging (Pool 1)      : {audit_stats['L2_Local_Kriging']}")
    print(f"Level 3 - Climatological Bilinear     : {audit_stats['L3_Climatology']}")
    print("=" * 55)


if __name__ == "__main__":
    run_spm_imputation()