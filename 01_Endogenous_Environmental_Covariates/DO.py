import pandas as pd
import numpy as np
import os
import glob
import warnings
import xarray as xr
from sklearn.neighbors import BallTree
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    from pykrige.ok import OrdinaryKriging
except ImportError:
    pass

warnings.filterwarnings('ignore')

# ================= 1. Global Configuration & QC Parameters =================
INPUT_FILE = 'DATA.xlsx'
OUTPUT_FILE = 'DATA_DO_Imputed_Final.xlsx'

NC_DIR = r"C:\Users\Administrator\PycharmProjects\PythonProject3\CEMES_DO"

# Dual-tolerance Quality Control (QC)
RATIO_THRESHOLD = 2.0
ABS_TOLERANCE = 3.0
O2_MOLAR_MASS_CONVERSION = 0.032

SEARCH_RADIUS_KRIGING = 0.5

# Haversine Conversion Constants for Geostatistical Modeling
EARTH_RADIUS_KM = 6371.01
DEG_TO_KM = 111.32
SEARCH_RADIUS_RAD = (SEARCH_RADIUS_KRIGING * DEG_TO_KM) / EARTH_RADIUS_KM

COASTAL_PADDING_LIMIT = 1  # 1 pixel padding (~27km) for coastal extrapolation

# Seamless Regional Geo-provinces
SEA_BOUNDARIES = {
    'Bohai': {'lat': (37.0, 42.0), 'lon': (117.0, 122.5)},
    'Yellow': {'lat': (31.0, 37.0), 'lon': (118.0, 127.0)},
    'East': {'lat': (23.0, 31.0), 'lon': (116.0, 131.0)},
    'South': {'lat': (3.0, 23.0), 'lon': (105.0, 121.0)}
}


# ================= 2. Core Methodologies =================

def get_region(lon, lat):
    for sea, bounds in SEA_BOUNDARIES.items():
        if bounds['lat'][0] <= lat <= bounds['lat'][1] and \
                bounds['lon'][0] <= lon <= bounds['lon'][1]:
            return sea
    return 'OpenOcean'


def extract_cmems_do_bilinear(ds, lat, lon, time):
    """
    Tier 1: DO extraction featuring Bilinear Downscaling for 0.25° to sub-grid mapping.
    Includes coastal padding to prevent NaN mask blindness.
    """
    try:
        ds_time = ds['o2'].sel(time=time, method='nearest')

        # Extrapolate marine data to coastal margins to prevent interpolation failure
        da_padded = ds_time.ffill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .ffill(dim='latitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='latitude', limit=COASTAL_PADDING_LIMIT)

        # Bilinear interpolation for smooth sub-grid spatial gradient recovery
        target_val = float(
            da_padded.interp(latitude=lat, longitude=lon, method='linear').values) * O2_MOLAR_MASS_CONVERSION

        if np.isnan(target_val):
            return np.nan, "Land_Mask_Failed"

        return target_val, "Valid"
    except Exception:
        return np.nan, "Extraction_Error"


# ================= 3. Three-Stage Sequential Assimilation Pipeline =================

def run_do_imputation():
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(f"[CRITICAL ERROR] Missing source file: {INPUT_FILE}")

    nc_files = glob.glob(os.path.join(NC_DIR, "*.nc"))
    if not nc_files:
        print(f"[WARNING] No .nc files found in {NC_DIR}. Proceeding with L2/L3 only.")
        ds_bgc = None
    else:
        try:
            ds_bgc = xr.open_mfdataset(nc_files, combine='by_coords').isel(depth=0)
        except Exception as e:
            print(f"[WARNING] NetCDF structural load failed: {e}. Proceeding with L2/L3 only.")
            ds_bgc = None

    df = pd.read_excel(INPUT_FILE)
    df.columns = df.columns.str.strip()

    col_do = 'DO'
    col_source = 'DO_Source'

    # [CRITICAL] State Initialization: Strict alignment with Original Data
    if col_source not in df.columns:
        df.insert(df.columns.get_loc(col_do) + 1, col_source, np.nan)

    mask_original = df[col_do].notna() & (df.get(col_source, 'Unfilled') != 'CMEMS_Bilinear_Downscaled') & (
                df.get(col_source, 'Unfilled') != 'Kriging_Interpolation') & (
                                df.get(col_source, 'Unfilled') != 'Regional_Median')

    df[col_source] = 'Unfilled'
    df.loc[mask_original, col_source] = 'Original Data'

    audit_stats = {'L1_Success': 0, 'L1_Masked': 0, 'L2_Kriging': 0, 'L3_Median': 0}

    # --- STAGE 1: CMEMS Extraction (L1) ---
    print("\n--- STAGE 1: Executing CMEMS Reanalysis Bilinear Downscaling ---")
    if ds_bgc is not None:
        for idx in df[df[col_source] == 'Unfilled'].index:
            lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
            time = pd.to_datetime(df.at[idx, 'Samplingtime'])

            val_l1, status = extract_cmems_do_bilinear(ds_bgc, lat, lon, time)
            if status == "Valid" and pd.notna(val_l1):
                df.at[idx, col_do] = val_l1
                df.at[idx, col_source] = "CMEMS_Bilinear_Downscaled"
                audit_stats['L1_Success'] += 1
            else:
                audit_stats['L1_Masked'] += 1

    # --- ASSIMILATION 1: Build Pool_1 for Kriging ---
    pool_1 = df[df[col_do].notna()]
    # Enforce [Lat, Lon] exact order for Haversine logic
    ref_coords_raw = pool_1[['Latitude', 'Longitude']].values
    ref_coords_rad = np.radians(ref_coords_raw)
    ref_values_1 = pool_1[col_do].values

    # --- STAGE 2: Constrained Ordinary Kriging (L2) ---
    print("--- STAGE 2: Executing Ordinary Kriging (Based on Assimilated Pool 1) ---")
    unfilled_after_l1 = df[df[col_source] == 'Unfilled'].index

    if len(ref_coords_rad) >= 3:
        tree = BallTree(ref_coords_rad, metric='haversine')

        for idx in unfilled_after_l1:
            lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
            query_rad = np.radians([[lat, lon]])

            nb_idx = tree.query_radius(query_rad, r=SEARCH_RADIUS_RAD)[0]

            if len(nb_idx) >= 3:
                try:
                    # PyKrige requires [x, y] -> [Longitude, Latitude]
                    ok = OrdinaryKriging(
                        ref_coords_raw[nb_idx, 1], ref_coords_raw[nb_idx, 0], ref_values_1[nb_idx],
                        variogram_model='linear', verbose=False, enable_plotting=False,
                        coordinates_type='geographic'
                    )
                    z, _ = ok.execute('grid', [lon], [lat])
                    df.at[idx, col_do] = max(0.0, float(z.data[0][0]))
                    df.at[idx, col_source] = "Kriging_Interpolation"
                    audit_stats['L2_Kriging'] += 1
                except Exception:
                    pass

    # --- ASSIMILATION 2: Build Pool_2 for Regional Median ---
    pool_2 = df[df[col_do].notna()]
    ref_values_2 = pool_2[col_do].values

    # --- STAGE 3: Dynamic Regional Median Fallback (L3) ---
    print("--- STAGE 3: Executing Regional Median Fallback (Based on Fully Assimilated Pool 2) ---")
    unfilled_after_l2 = df[df[col_source] == 'Unfilled'].index

    for idx in unfilled_after_l2:
        lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
        region = get_region(lon, lat)

        region_mask = pool_2.apply(lambda x: get_region(x['Longitude'], x['Latitude']), axis=1) == region
        reg_median = pool_2[region_mask][col_do].median()

        final_val = reg_median if pd.notna(reg_median) else np.median(ref_values_2)
        df.at[idx, col_do] = max(0.0, float(final_val)) if pd.notna(final_val) else 0.0
        df.at[idx, col_source] = "Regional_Median"
        audit_stats['L3_Median'] += 1

    # ================= 4. Data Export & Visual Formatting =================
    df.to_excel(OUTPUT_FILE, index=False)

    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        red_font = Font(color="FF0000")

        header_map = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
        if col_do in header_map and col_source in header_map:
            t_idx, s_idx = header_map[col_do], header_map[col_source]
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=s_idx).value != "Original Data":
                    ws.cell(row=r, column=t_idx).font = red_font
                    ws.cell(row=r, column=t_idx).number_format = '0.00'

        wb.save(OUTPUT_FILE)
    except Exception:
        pass

    # ================= 5. Academic Audit Report =================
    print("\n" + "=" * 55)
    print("      DO BILINEAR DOWNSCALING ASSIMILATION AUDIT")
    print("=" * 55)
    print(f"Level 1 - CMEMS Bilinear Downscaled      : {audit_stats['L1_Success']}")
    print(f"Level 1 - Land Mask / Missing Data       : {audit_stats['L1_Masked']}")
    print(f"Level 2 - Ordinary Kriging (Dynamic Pool): {audit_stats['L2_Kriging']}")
    print(f"Level 3 - Regional Median (Dynamic Pool) : {audit_stats['L3_Median']}")
    print("=" * 55)


if __name__ == "__main__":
    run_do_imputation()