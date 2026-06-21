import pandas as pd
import numpy as np
from sklearn.neighbors import BallTree, KNeighborsRegressor
import warnings
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    from pykrige.ok import OrdinaryKriging
except ImportError:
    raise ImportError("[CRITICAL] pykrige library required.")

warnings.filterwarnings('ignore')

# ================= 1. Configuration & Boundaries =================
SEA_BOUNDARIES = {
    'Bohai': {'lat': (37.0, 42.0), 'lon': (117.0, 122.5)},
    'Yellow': {'lat': (31.0, 37.0), 'lon': (119.0, 127.0)},
    'East': {'lat': (23.0, 31.0), 'lon': (117.0, 131.0)},
    'South': {'lat': (3.0, 23.0), 'lon': (105.0, 121.0)}
}

INPUT_TARGET_FILE = 'DATA.xlsx'
INPUT_REF_FILE = 'Sediment TOC.xlsx'
OUTPUT_FILE = 'TOC_Imputed_Result.xlsx'

SEARCH_RADIUS_DEG = 0.5
KNN_NEIGHBORS = 5

# Haversine Conversion Constants
EARTH_RADIUS_KM = 6371.01
DEG_TO_KM = 111.32
SEARCH_RADIUS_RAD = (SEARCH_RADIUS_DEG * DEG_TO_KM) / EARTH_RADIUS_KM


# ================= 2. Core Modules =================

def identify_region(lon, lat):
    for sea, bounds in SEA_BOUNDARIES.items():
        if bounds['lat'][0] <= lat <= bounds['lat'][1] and \
                bounds['lon'][0] <= lon <= bounds['lon'][1]:
            return sea
    return 'OpenOcean'


def run_toc_imputation():
    if not os.path.exists(INPUT_TARGET_FILE) or not os.path.exists(INPUT_REF_FILE):
        raise FileNotFoundError("Missing input data files.")

    df_target = pd.read_excel(INPUT_TARGET_FILE)
    df_target.columns = df_target.columns.str.strip()

    df_ref = pd.read_excel(INPUT_REF_FILE)
    df_ref.columns = df_ref.columns.str.strip()

    col_toc = 'TOC (%)'
    col_source = 'TOC_Source'

    if col_source not in df_target.columns:
        df_target[col_source] = df_target[col_toc].apply(lambda x: 'Original Data' if pd.notna(x) else 'Unfilled')


    df_target['Temp_Region'] = df_target.apply(lambda row: identify_region(row['Longitude'], row['Latitude']), axis=1)

    imputed_registry = []
    audit_stats = {'L1': 0, 'L2': 0, 'L3': 0}

    unfilled_idx = df_target[df_target[col_source] == 'Unfilled'].index

    # --- PHASE 1: Extended Ordinary Kriging (L1) ---
    pool_ext = df_ref[df_ref[col_toc].notna()].copy()
    if not pool_ext.empty:
        ref_lat_col = next((c for c in pool_ext.columns if c.lower() in ['latitude', 'lat']), None)
        ref_lon_col = next((c for c in pool_ext.columns if c.lower() in ['longitude', 'lon']), None)

        if ref_lat_col and ref_lon_col:
            # Enforce [Lat, Lon] exact order
            ref_coords_ext = pool_ext[[ref_lat_col, ref_lon_col]].values
            ref_coords_rad = np.radians(ref_coords_ext)
            ref_vals = pool_ext[col_toc].values

            tree = BallTree(ref_coords_rad, metric='haversine')

            for idx in unfilled_idx:
                lat, lon = df_target.at[idx, 'Latitude'], df_target.at[idx, 'Longitude']
                if pd.isna(lat) or pd.isna(lon): continue

                query_rad = np.radians([[lat, lon]])
                nb_indices = tree.query_radius(query_rad, r=SEARCH_RADIUS_RAD)[0]

                if len(nb_indices) >= 3:
                    try:
                        # Enforce geographic coordinates for Kriging
                        ok = OrdinaryKriging(
                            ref_coords_ext[nb_indices, 1], ref_coords_ext[nb_indices, 0], ref_vals[nb_indices],
                            variogram_model='linear', verbose=False, coordinates_type='geographic'
                        )
                        z, _ = ok.execute('grid', [lon], [lat])
                        val_l1 = float(z.data[0][0])

                        if val_l1 > 0:
                            df_target.at[idx, col_toc] = val_l1
                            df_target.at[idx, col_source] = 'Extended_Kriging'
                            audit_stats['L1'] += 1
                            imputed_registry.append(idx)
                    except (ValueError, np.linalg.LinAlgError):
                        pass

    # --- PHASE 2: Spatial KNN (L2) ---
    unfilled_idx_phase2 = df_target[df_target[col_source] == 'Unfilled'].index
    pool_1 = df_target[df_target[col_source] != 'Unfilled'].copy()

    if not pool_1.empty and len(unfilled_idx_phase2) > 0:
        ref_coords_1 = pool_1[['Latitude', 'Longitude']].values
        ref_coords_rad = np.radians(ref_coords_1)
        ref_vals = pool_1[col_toc].values

        tree = BallTree(ref_coords_rad, metric='haversine')

        for idx in unfilled_idx_phase2:
            lat, lon = df_target.at[idx, 'Latitude'], df_target.at[idx, 'Longitude']
            if pd.isna(lat) or pd.isna(lon): continue

            query_rad = np.radians([[lat, lon]])
            nb_indices = tree.query_radius(query_rad, r=SEARCH_RADIUS_RAD)[0]

            if len(nb_indices) >= 3:
                knn = KNeighborsRegressor(
                    n_neighbors=min(KNN_NEIGHBORS, len(nb_indices)),
                    weights='distance',
                    algorithm='ball_tree',
                    metric='haversine'
                )
                knn.fit(ref_coords_rad[nb_indices], ref_vals[nb_indices])
                val_l2 = float(knn.predict(query_rad)[0])

                if val_l2 > 0:
                    df_target.at[idx, col_toc] = val_l2
                    df_target.at[idx, col_source] = 'Spatial_KNN'
                    audit_stats['L2'] += 1
                    imputed_registry.append(idx)

    # --- PHASE 3: Regional Median (L3) ---
    unfilled_idx_phase3 = df_target[df_target[col_source] == 'Unfilled'].index
    pool_2 = df_target[df_target[col_source] != 'Unfilled'].copy()

    for idx in unfilled_idx_phase3:
        region = df_target.at[idx, 'Temp_Region']
        region_mask = pool_2['Temp_Region'] == region

        reg_median = pool_2.loc[region_mask, col_toc].median() if region_mask.any() else np.nan
        final_val = reg_median if pd.notna(reg_median) else pool_2[col_toc].median()

        df_target.at[idx, col_toc] = max(0.0, float(final_val)) if pd.notna(final_val) else 0.0
        df_target.at[idx, col_source] = 'Regional_Median'
        audit_stats['L3'] += 1
        imputed_registry.append(idx)

    # Clean tracking columns
    df_target.drop(columns=['Temp_Region'], inplace=True, errors='ignore')

    # ================= 4. Data Export & Visual Formatting =================
    df_target.to_excel(OUTPUT_FILE, index=False)

    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        red_font = Font(color="FF0000")

        header_map = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
        if col_toc in header_map and col_source in header_map:
            c_idx = header_map[col_toc]
            s_idx = header_map[col_source]
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=s_idx).value != "Original Data":
                    ws.cell(row=r, column=c_idx).font = red_font
                    ws.cell(row=r, column=c_idx).number_format = '0.00'

        wb.save(OUTPUT_FILE)
    except Exception as e:
        pass

    # ================= 5. Academic Audit Report =================
    print("\n" + "=" * 55)
    print("      SEDIMENT TOC IMPUTATION AUDIT")
    print("=" * 55)
    print(f"Level 1 - Extended Kriging (Haversine) : {audit_stats['L1']}")
    print(f"Level 2 - Spatial KNN (Haversine)      : {audit_stats['L2']}")
    print(f"Level 3 - Regional/Global Median       : {audit_stats['L3']}")
    print("=" * 55)


if __name__ == "__main__":
    run_toc_imputation()