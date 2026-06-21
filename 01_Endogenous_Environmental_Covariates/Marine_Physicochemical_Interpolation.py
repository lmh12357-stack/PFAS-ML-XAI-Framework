import pandas as pd
import numpy as np
import os
import warnings
import xarray as xr
from sklearn.neighbors import BallTree
from openpyxl import load_workbook
from openpyxl.styles import Font

try:
    from pykrige.ok import OrdinaryKriging
except ImportError:
    raise ImportError("Critical dependency missing: pykrige.")

warnings.filterwarnings('ignore')

# ================= 1. Global Academic Configuration =================
INPUT_FILE = 'DATA.xlsx'
OUTPUT_FILE = 'Physicochemical_Imputed_Result.xlsx'

NC_PATH_PHY = 'cmems_mod_glo_phy_my_0.083deg_P1M-m_1767061731823.nc'
NC_PATH_BGC = 'cmems_mod_glo_bgc_my_0.25deg_P1M-m_1766974042700.nc'

EARTH_RADIUS_KM = 6371.01
DEG_TO_KM = 111.32

# Kriging Search Radius (~55.66 km) based on spherical distance
SEARCH_RADIUS_L2_RAD = (0.50 * DEG_TO_KM) / EARTH_RADIUS_KM

# Coastal Padding Limit for Track A & C (preventing NaN blindness)
COASTAL_PADDING_LIMIT = 2

# Track B (Salinity) L1.5 Estuarine Prior Constraints
ESTUARINE_SALINITY_THRESHOLD = 10.0
OFFSHORE_PULL_THRESHOLD = 25.0

VAR_CONFIG = {
    'Temperature (°C)': {
        'nc_source': 'PHY', 'nc_var': 'thetao', 'constraints': [-2.0, 45.0],
        'l1_method': 'nearest_padded'
    },
    'Salinity (‰)': {
        'nc_source': 'PHY', 'nc_var': 'so', 'constraints': [0.0, 42.0],
        'l1_method': 'nearest_unpadded'
    },
    'pH': {
        'nc_source': 'BGC', 'nc_var': 'ph', 'constraints': [6.0, 9.0],
        'l1_method': 'bilinear_padded'
    }
}


# ================= 2. Core Extractor Engines =================

def extract_cmems_data(ds, nc_var, lat, lon, time_val, method):
    try:
        da_time = ds[nc_var].sel(time=time_val, method='nearest')

        # Track B: Pure nearest extraction, zero padding to prevent offshore contamination
        if method == 'nearest_unpadded':
            val = float(da_time.sel(latitude=lat, longitude=lon, method='nearest').values)
            return val if pd.notna(val) else np.nan

        # Track A & C: Sea-over-Land Extrapolation to resolve spatial dead zones
        da_padded = da_time.ffill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .ffill(dim='latitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='latitude', limit=COASTAL_PADDING_LIMIT)

        if method == 'nearest_padded':
            val = float(da_padded.sel(latitude=lat, longitude=lon, method='nearest').values)
        elif method == 'bilinear_padded':
            val = float(da_padded.interp(latitude=lat, longitude=lon, method='linear').values)
        else:
            return np.nan

        return val if pd.notna(val) else np.nan
    except (KeyError, ValueError, IndexError):
        return np.nan


def extract_l3_climatology(ds_mean, nc_var, lat, lon, month_val):
    try:
        da_month = ds_mean[nc_var].sel(month=month_val)
        da_padded = da_month.ffill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='longitude', limit=COASTAL_PADDING_LIMIT) \
            .ffill(dim='latitude', limit=COASTAL_PADDING_LIMIT) \
            .bfill(dim='latitude', limit=COASTAL_PADDING_LIMIT)
        val = float(da_padded.sel(latitude=lat, longitude=lon, method='nearest').values)
        return val if pd.notna(val) else np.nan
    except (KeyError, ValueError, IndexError):
        return np.nan


# ================= 3. Assimilation Pipeline =================

def run_interpolation_pipeline():
    if not os.path.exists(INPUT_FILE):
        raise FileNotFoundError(f"Missing input file: {INPUT_FILE}")

    ds_phy = xr.open_dataset(NC_PATH_PHY).isel(depth=0)
    ds_phy_mean = ds_phy.groupby('time.month').mean('time').load()

    ds_bgc = xr.open_dataset(NC_PATH_BGC).isel(depth=0)
    ds_bgc_mean = ds_bgc.groupby('time.month').mean('time').load()

    df = pd.read_excel(INPUT_FILE)
    df.columns = df.columns.str.strip()

    df['Temp_Time'] = pd.to_datetime(df['Samplingtime'].astype(str).str.replace('.', '-'), errors='coerce')
    df['Temp_YM'] = df['Temp_Time'].dt.to_period('M')
    df['Temp_Month'] = df['Temp_Time'].dt.month

    # Pre-compute Estuarine Prior Matrix for Track B topological isolation
    estuarine_prior = {}
    if 'Salinity (‰)' in df.columns:
        sal_orig = df[df['Salinity (‰)'].notna() & (df.get('Salinity (‰)_Source', 'Original Data') == 'Original Data')]
        if not sal_orig.empty:
            hist_stats = sal_orig.groupby(['Latitude', 'Longitude'])['Salinity (‰)'].agg(['min', 'median'])
            estuarine_prior = hist_stats.to_dict(orient='index')

    imputed_registry = []

    # Granular Audit Matrix
    audit_stats = {
        var: {
            'L1': 0, 'L1.5': 0,
            'L2_Attempted': 0, 'L2_Success': 0, 'L2_Crash': 0, 'L2_Rejected': 0,
            'L3': 0
        } for var in VAR_CONFIG
    }

    for var_name, cfg in VAR_CONFIG.items():
        if var_name not in df.columns: continue

        source_col = f"{var_name}_Source"
        df[source_col] = df[var_name].apply(lambda x: 'Original Data' if pd.notna(x) else 'Unfilled')

        ds = ds_phy if cfg['nc_source'] == 'PHY' else ds_bgc
        ds_mean = ds_phy_mean if cfg['nc_source'] == 'PHY' else ds_bgc_mean

        isolated_estuary_registry = set()

        # --- L1 & L1.5: Satellite Assimilation & Estuarine Relaxation ---
        unfilled_idx = df[df[source_col] == 'Unfilled'].index
        for idx in unfilled_idx:
            lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
            time_val = df.at[idx, 'Temp_Time']

            if pd.isna(lat) or pd.isna(lon) or pd.isna(time_val): continue

            val_l1 = extract_cmems_data(ds, cfg['nc_var'], lat, lon, time_val, cfg['l1_method'])
            val_assigned = np.nan
            source_assigned = ""

            if var_name == 'Salinity (‰)':
                coord = (lat, lon)
                is_estuary = (coord in estuarine_prior) and (
                            estuarine_prior[coord]['min'] < ESTUARINE_SALINITY_THRESHOLD)

                if is_estuary:
                    isolated_estuary_registry.add(idx)
                    if pd.isna(val_l1) or val_l1 >= OFFSHORE_PULL_THRESHOLD:
                        val_assigned = estuarine_prior[coord]['median']
                        source_assigned = "L1.5_Estuarine_Temporal_Relaxation"
                    else:
                        val_assigned = val_l1
                        source_assigned = "L1_CMEMS_Reanalysis"
                else:
                    val_assigned = val_l1
                    source_assigned = "L1_CMEMS_Reanalysis"
            else:
                val_assigned = val_l1
                if cfg['l1_method'] == 'bilinear_padded':
                    source_assigned = "L1_CMEMS_Bilinear_Downscaled"
                else:
                    source_assigned = "L1_CMEMS_Reanalysis"

            if pd.notna(val_assigned) and cfg['constraints'][0] <= val_assigned <= cfg['constraints'][1]:
                df.at[idx, var_name] = val_assigned
                df.at[idx, source_col] = source_assigned
                if "L1.5" in source_assigned:
                    audit_stats[var_name]['L1.5'] += 1
                else:
                    audit_stats[var_name]['L1'] += 1
                imputed_registry.append((idx, var_name))

        # --- L2: Augmented Spatiotemporally Isolated Kriging ---
        grouped_ym = df.dropna(subset=['Temp_YM']).groupby('Temp_YM')
        for ym, group in grouped_ym:
            # Pool Augmented with CMEMS reanalysis data to anchor local spatial interpolation
            pool_mask = group[var_name].notna() & group[source_col].str.contains('Original|CMEMS', na=False)
            pool_data = group[pool_mask].copy()

            if len(pool_data) >= 3:
                agg_pool = pool_data.groupby(['Latitude', 'Longitude'], as_index=False)[var_name].mean()
                ref_rad = np.radians(agg_pool[['Latitude', 'Longitude']].values)
                ref_deg = agg_pool[['Longitude', 'Latitude']].values
                ref_val = agg_pool[var_name].values

                if len(ref_val) < 3: continue
                spatial_tree = BallTree(ref_rad, metric='haversine')

                unfilled_group_idx = group[group[source_col] == 'Unfilled'].index
                for idx in unfilled_group_idx:
                    if idx in isolated_estuary_registry:
                        continue

                    lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
                    query_rad = np.radians([[lat, lon]])

                    nb_idx = spatial_tree.query_radius(query_rad, r=SEARCH_RADIUS_L2_RAD)[0]
                    if len(nb_idx) >= 3:
                        audit_stats[var_name]['L2_Attempted'] += 1
                        try:
                            # Enforce spherical geostatistical modeling
                            ok = OrdinaryKriging(
                                ref_deg[nb_idx, 0], ref_deg[nb_idx, 1], ref_val[nb_idx],
                                variogram_model='linear', verbose=False, coordinates_type='geographic'
                            )
                            z, _ = ok.execute('grid', [lon], [lat])
                            val_l2 = float(z.data[0][0])

                            if cfg['constraints'][0] <= val_l2 <= cfg['constraints'][1]:
                                df.at[idx, var_name] = val_l2
                                df.at[idx, source_col] = "L2_Isolated_Kriging"
                                audit_stats[var_name]['L2_Success'] += 1
                                imputed_registry.append((idx, var_name))
                            else:
                                print(
                                    f"[L2 Rejection] {var_name} | Lat:{lat:.4f}, Lon:{lon:.4f} | Computed: {val_l2:.4f} | Out of bounds.")
                                audit_stats[var_name]['L2_Rejected'] += 1

                        except (ValueError, np.linalg.LinAlgError) as e:
                            print(f"[L2 Matrix Crash] {var_name} | Lat:{lat:.4f}, Lon:{lon:.4f} | Exception: {e}")
                            audit_stats[var_name]['L2_Crash'] += 1

        # --- L3: Unconstrained Climatology Fallback ---
        remaining_idx = df[df[source_col] == 'Unfilled'].index
        for idx in remaining_idx:
            if idx in isolated_estuary_registry:
                continue

            lat, lon = df.at[idx, 'Latitude'], df.at[idx, 'Longitude']
            month_val = df.at[idx, 'Temp_Month']

            val_l3 = extract_l3_climatology(ds_mean, cfg['nc_var'], lat, lon, month_val)
            if pd.notna(val_l3) and cfg['constraints'][0] <= val_l3 <= cfg['constraints'][1]:
                df.at[idx, var_name] = val_l3
                df.at[idx, source_col] = "L3_Climatological_Fallback"
                audit_stats[var_name]['L3'] += 1
                imputed_registry.append((idx, var_name))

    # ================= 4. Export & Formatting =================
    df.drop(columns=['Temp_Time', 'Temp_YM', 'Temp_Month'], inplace=True, errors='ignore')
    df.to_excel(OUTPUT_FILE, index=False)

    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        red_font = Font(color="FF0000")
        headers = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
        for r_idx, v_name in imputed_registry:
            if v_name in headers:
                ws.cell(row=r_idx + 2, column=headers[v_name]).font = red_font
        wb.save(OUTPUT_FILE)
    except Exception as e:
        pass

    # ================= 5. High-Resolution Audit Print =================
    print("\n" + "=" * 65)
    print("      Physicochemical Assimilation & Transparency Audit")
    print("=" * 65)
    for var, stats in audit_stats.items():
        print(f"[{var}]")
        print(f"  L1   (Satellite/Downscaled) : {stats['L1']}")
        if var == 'Salinity (‰)':
            print(f"  L1.5 (Estuarine Relaxation) : {stats['L1.5']}")
        print(
            f"  L2   (Kriging Attempted)    : {stats['L2_Attempted']} (Success: {stats['L2_Success']} | Crash: {stats['L2_Crash']} | Rejected: {stats['L2_Rejected']})")
        print(f"  L3   (Climatology Fallback) : {stats['L3']}")
        print("-" * 65)


if __name__ == "__main__":
    run_interpolation_pipeline()